from __future__ import annotations

import asyncio
from collections.abc import AsyncIterator, Generator, Sequence
from pathlib import Path
from typing import Any, cast

import pandas as pd
import structlog

from awa_common.metrics import record_etl_normalize_error, record_etl_rows_normalized
from awa_common.settings import Settings
from awa_common.types import PriceRow as PriceRowDict, PriceRowModel
from awa_common.vendor import normalize_currency, normalize_sku, parse_decimal

from .normaliser import normalise
from .reader import detect_format

logger = structlog.get_logger(__name__)

SETTINGS = Settings()
DEFAULT_BATCH_SIZE = SETTINGS.PRICE_IMPORTER_CHUNK_ROWS
VALIDATION_WORKERS = SETTINGS.PRICE_IMPORTER_VALIDATION_WORKERS


def _parse_int(value: Any, *, default: int = 0) -> int:
    if value in (None, ""):
        return default
    try:
        # Handle pandas NA / floats by casting to float first.
        if isinstance(value, str):
            text = value.strip()
            if not text:
                return default
            return int(float(text))
        return int(float(value))
    except (TypeError, ValueError):
        return default


def validate_price_rows(rows: Sequence[dict[str, Any]]) -> list[PriceRowDict]:
    """Validate and normalise a batch of price rows."""
    valid: list[PriceRowDict] = []
    errors: list[dict[str, Any]] = []

    for idx, raw in enumerate(rows):
        try:
            sku = normalize_sku(raw.get("sku"))
            cost = parse_decimal(raw.get("cost"))
            if cost < 0:
                raise ValueError("cost negative")
            currency = normalize_currency(raw.get("currency"))
            moq = _parse_int(raw.get("moq"), default=0)
            lead_time = _parse_int(raw.get("lead_time_days"), default=0)
            payload = PriceRowModel(
                sku=sku,
                unit_price=cost,
                currency=currency,
                moq=moq,
                lead_time_d=lead_time,
            ).model_dump()
            valid.append(cast(PriceRowDict, payload))
        except ValueError as exc:
            errors.append({"index": idx, "error": str(exc), "row": dict(raw)})

    if errors:
        record_etl_normalize_error("price_import", "row_validation", len(errors))
        sample = errors[:3]
        raise ValueError(
            f"{len(errors)} invalid price rows; sample={sample}",
        )
    record_etl_rows_normalized("price_import", len(valid))
    return valid


def _iter_csv_chunks(path: str | Path, batch_size: int) -> Generator[pd.DataFrame]:
    """Yield CSV chunks while retrying encodings and delimiters like the legacy reader."""
    encodings = ("utf-8", "utf-8-sig", "cp1252")
    for enc in encodings:
        try:
            reader = pd.read_csv(
                path,
                sep=None,
                engine="python",
                encoding=enc,
                chunksize=batch_size,
            )
            yield from reader
            return
        except UnicodeDecodeError:
            continue
        except Exception:
            continue

    for sep in (",", ";", "\t", "|"):
        try:
            reader = pd.read_csv(path, sep=sep, chunksize=batch_size)
            yield from reader
            return
        except Exception:
            continue
    raise RuntimeError(f"Failed to stream CSV file: {path}")


def _iter_xlsx_chunks(path: str | Path, batch_size: int) -> Generator[pd.DataFrame]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - optional dependency guard
        raise RuntimeError("openpyxl is required to stream XLSX files") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        headers: list[str] | None = None
        buffer: list[list[Any]] = []
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                if not row:
                    continue
                headers = [str(cell).strip() if cell is not None else "" for cell in row]
                continue
            if row is None or not any(value is not None for value in row):
                continue
            normalized_row = list(row[: len(headers)])
            if len(normalized_row) < len(headers):
                normalized_row.extend([None] * (len(headers) - len(normalized_row)))
            buffer.append(normalized_row)
            if len(buffer) >= batch_size:
                yield pd.DataFrame(buffer, columns=headers)
                buffer.clear()
        if headers and buffer:
            yield pd.DataFrame(buffer, columns=headers)
    finally:
        wb.close()


def _frame_iterator(path: str | Path, batch_size: int) -> Generator[pd.DataFrame]:
    fmt = detect_format(path)
    if fmt == "csv":
        return _iter_csv_chunks(path, batch_size)
    if fmt == "excel":
        return _iter_xlsx_chunks(path, batch_size)
    raise RuntimeError(f"Unsupported price importer format: {fmt}")


def _next_frame(iterator: Generator[pd.DataFrame]) -> pd.DataFrame | None:
    try:
        return next(iterator)
    except StopIteration:
        return None


async def _validate_frame(frame: pd.DataFrame, sem: asyncio.Semaphore) -> list[PriceRowDict]:
    async with sem:
        return await asyncio.to_thread(_normalize_and_validate, frame)


def _normalize_and_validate(frame: pd.DataFrame) -> list[PriceRowDict]:
    cleaned = normalise(frame)
    records = cleaned.to_dict(orient="records")
    if not records:
        return []
    return validate_price_rows(records)


async def iter_price_batches(
    path: str | Path,
    batch_size: int | None = None,
    max_workers: int | None = None,
) -> AsyncIterator[list[PriceRowDict]]:
    """
    Yield validated, normalised price rows with bounded concurrency.

    The caller must iterate using ``async for`` to benefit from streaming behaviour.
    """

    target_batch = max(1, int(batch_size or DEFAULT_BATCH_SIZE))
    worker_limit = max(1, int(max_workers or VALIDATION_WORKERS))
    iterator = _frame_iterator(path, target_batch)
    sem = asyncio.Semaphore(worker_limit)
    pending: dict[int, asyncio.Task[list[PriceRowDict]]] = {}
    next_yield = 0
    idx = 0

    try:
        while True:
            frame = await asyncio.to_thread(_next_frame, iterator)
            if frame is None:
                break
            if frame.empty:
                continue
            task = asyncio.create_task(_validate_frame(frame, sem))
            pending[idx] = task
            idx += 1
            if len(pending) >= worker_limit:
                batch = await _await_ordered(pending, next_yield)
                if batch:
                    yield batch
                next_yield += 1

        while pending:
            batch = await _await_ordered(pending, next_yield)
            if batch:
                yield batch
            next_yield += 1
    finally:
        for task in pending.values():
            task.cancel()


async def _await_ordered(pending: dict[int, asyncio.Task[list[PriceRowDict]]], order: int) -> list[PriceRowDict]:
    task = pending.pop(order, None)
    if task is None:
        raise RuntimeError("price importer batch ordering mismatch")
    result = await task
    if not result:
        return []
    return result
