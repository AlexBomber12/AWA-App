from __future__ import annotations

import asyncio
import csv
import io
from dataclasses import asdict, dataclass
from datetime import date, datetime
from decimal import Decimal
from typing import Any
from urllib.parse import parse_qs, urlparse

import anyio
import structlog

from awa_common.http_client import AsyncHTTPClient
from awa_common.metrics import record_etl_normalize_error, record_etl_rows_normalized
from awa_common.minio import create_boto3_client, get_s3_client_config
from awa_common.retries import RetryConfig, aretry
from awa_common.settings import Settings
from awa_common.types import RateRowModel
from awa_common.vendor import parse_date as vendor_parse_date, parse_decimal

URL = Settings().FREIGHT_API_URL

__all__ = [
    "URL",
    "UnsupportedExcelError",
    "UnsupportedFileFormatError",
    "fetch_rates",
    "fetch_sources",
]

logger = structlog.get_logger(__name__).bind(component="logistics_etl.client")
_HTTP_CLIENT: AsyncHTTPClient | None = None
_HTTP_CLIENT_CONFIG: tuple[float, int] | None = None
_HTTP_LOCK = asyncio.Lock()


class UnsupportedExcelError(RuntimeError):
    """Raised when an Excel payload cannot be parsed due to missing optional deps."""


class UnsupportedFileFormatError(RuntimeError):
    """Raised when the payload cannot be mapped to CSV/XLS/XLSX/ZIP formats."""


@dataclass
class SourcePayload:
    source: str
    raw: bytes
    meta: dict[str, Any]
    rows: list[dict[str, Any]]
    error: Exception | None = None


async def fetch_sources() -> list[dict[str, Any]]:
    """
    Fetch logistics rate sources configured via LOGISTICS_SOURCES.

    Returns a list of dictionaries each containing:
        source, raw (bytes), meta (dict), rows (normalized list), error (Exception|None)
    """
    cfg = Settings()
    sources_env = cfg.LOGISTICS_SOURCES or ""
    uris = [s.strip() for s in sources_env.split(",") if s.strip()]
    if not uris:
        return []

    timeout_s = int(cfg.LOGISTICS_TIMEOUT_S)
    retries = int(cfg.LOGISTICS_RETRIES)
    snapshots: list[SourcePayload] = []

    for uri in uris:
        try:
            raw, meta = await _download_with_retries(uri, timeout_s=timeout_s, retries=retries)
            rows = _parse_rows(uri, raw, meta)
            snapshots.append(SourcePayload(uri, raw, meta, rows))
        except UnsupportedExcelError as exc:
            logger.warning("Excel support unavailable for %s: %s", uri, exc)
            snapshots.append(SourcePayload(uri, b"", {}, [], exc))
        except UnsupportedFileFormatError as exc:
            logger.warning("Unsupported logistics payload for %s: %s", uri, exc)
            raise
        except Exception as exc:  # pragma: no cover - logged for observability
            logger.exception("Failed to fetch logistics source %s", uri)
            snapshots.append(SourcePayload(uri, b"", {}, [], exc))

    return [asdict(snap) for snap in snapshots]


async def fetch_rates() -> list[dict[str, object]]:
    """
    Backward-compatible shim used by legacy callers.
    Downloads the single FREIGHT_API_URL CSV and returns parsed rows.
    """
    cfg = Settings()
    url = cfg.FREIGHT_API_URL
    if not url:
        return []

    try:
        raw, _ = await _download_with_retries(
            url,
            timeout_s=int(cfg.LOGISTICS_TIMEOUT_S),
            retries=int(cfg.LOGISTICS_RETRIES),
        )
    except Exception:
        return []

    if not raw:
        return []

    try:
        text_content = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text_content = raw.decode("utf-8", errors="ignore")

    reader = csv.DictReader(io.StringIO(text_content))
    rows: list[dict[str, object]] = []
    for row in reader:
        value = row.get("eur_per_kg", 0)
        try:
            row["eur_per_kg"] = float(value)
        except (TypeError, ValueError):
            row["eur_per_kg"] = 0.0
        rows.append(dict(row))
    return rows


async def _ensure_http_client(timeout_s: float | None = None, retries: int | None = None) -> AsyncHTTPClient:
    global _HTTP_CLIENT, _HTTP_CLIENT_CONFIG
    cfg = Settings()
    fallback_timeout = getattr(cfg, "LOGISTICS_TIMEOUT_S", getattr(cfg, "HTTP_TOTAL_TIMEOUT_S", 60.0))
    desired_timeout = float(timeout_s if timeout_s is not None else fallback_timeout)
    desired_retries = max(
        1,
        int(retries if retries is not None else getattr(cfg, "LOGISTICS_RETRIES", 0)),
        int(getattr(cfg, "ETL_RETRY_ATTEMPTS", 1)),
        int(getattr(cfg, "HTTP_MAX_RETRIES", 1)),
    )
    desired = (desired_timeout, desired_retries)
    client = _HTTP_CLIENT
    if client is not None and _HTTP_CLIENT_CONFIG == desired:
        return client
    async with _HTTP_LOCK:
        client = _HTTP_CLIENT
        if client is not None and _HTTP_CLIENT_CONFIG == desired:
            return client
        if client is not None:
            await client.aclose()
        _HTTP_CLIENT = AsyncHTTPClient(
            integration="logistics_etl",
            total_timeout_s=desired_timeout,
            max_retries=desired_retries,
        )
        _HTTP_CLIENT_CONFIG = desired
        return _HTTP_CLIENT


async def _download_with_retries(url_or_uri: str, *, timeout_s: int, retries: int) -> tuple[bytes, dict[str, Any]]:
    parsed = urlparse(url_or_uri)
    scheme = (parsed.scheme or "").lower()
    cfg = Settings()
    attempts = max(1, int(getattr(cfg, "ETL_RETRY_ATTEMPTS", retries)), int(retries))
    raw_jitter = getattr(cfg, "ETL_RETRY_JITTER_S", 0.0)
    try:
        jitter_seconds = float(raw_jitter)
    except (TypeError, ValueError):
        jitter_seconds = 0.0
    retry_cfg = RetryConfig(
        operation="logistics_download",
        stop_after=attempts,
        base_wait_s=float(cfg.ETL_RETRY_BASE_S),
        max_wait_s=float(cfg.ETL_RETRY_MAX_S),
        jitter=jitter_seconds > 0,
        retry_on=(Exception,),
    )

    if scheme in ("http", "https"):
        return await _download_http(url_or_uri, timeout_s=timeout_s, retries=attempts)

    if scheme == "s3":

        async def target() -> tuple[bytes, dict[str, Any]]:
            return await _download_s3(parsed, timeout_s)
    elif scheme == "ftp":

        async def target() -> tuple[bytes, dict[str, Any]]:
            return await _download_ftp(parsed, timeout_s)
    else:
        raise ValueError(f"Unsupported logistics source scheme: {scheme or 'unknown'}")

    @aretry(retry_cfg)
    async def _download() -> tuple[bytes, dict[str, Any]]:
        return await target()

    return await _download()


class _LegacyHTTPWrapper:
    async def get_client(self) -> AsyncHTTPClient:
        return await _ensure_http_client()


http_client = _LegacyHTTPWrapper()


async def _download_http(
    url: str, timeout_s: int | None = None, retries: int | None = None
) -> tuple[bytes, dict[str, Any]]:
    client = await _ensure_http_client(timeout_s=timeout_s, retries=retries)
    response = await client.request("GET", url, follow_redirects=True, timeout=timeout_s)
    try:
        if hasattr(response, "raise_for_status"):
            response.raise_for_status()
        if hasattr(response, "aread"):
            body = await response.aread()
        else:
            body = getattr(response, "content", b"")
        etag = response.headers.get("etag")
        meta = {
            "content_type": response.headers.get("content-type"),
            "etag": etag.strip('"') if etag else None,
            "last_modified": response.headers.get("last-modified"),
        }
        seqno = response.headers.get("x-amz-version-id") or meta.get("etag")
        if seqno:
            meta["seqno"] = seqno.strip('"')
        return body or b"", meta
    finally:
        close = getattr(response, "aclose", None)
        if callable(close):
            await close()
        elif hasattr(response, "close"):
            response.close()


async def _download_s3(parsed, timeout_s: int) -> tuple[bytes, dict[str, Any]]:
    bucket = parsed.netloc
    key = parsed.path.lstrip("/")
    extra: dict[str, Any] = {}
    if parsed.query:
        params = parse_qs(parsed.query, keep_blank_values=True)
        version_id = params.get("versionId") or params.get("versionid")
        if version_id:
            extra["VersionId"] = version_id[0]

    def _get() -> tuple[bytes, dict[str, Any]]:
        client = create_boto3_client(
            config=get_s3_client_config(connect_timeout=timeout_s, read_timeout=timeout_s, max_pool_connections=None)
        )
        try:
            response = client.get_object(Bucket=bucket, Key=key, **extra)
        except Exception as exc:  # pragma: no cover - bubbled to retry
            raise RuntimeError(str(exc)) from exc
        body = response["Body"].read()
        meta = {
            "etag": (response.get("ETag") or "").strip('"') or None,
            "last_modified": None,
            "content_type": response.get("ContentType"),
            "seqno": response.get("VersionId"),
        }
        last_modified = response.get("LastModified")
        if last_modified:
            if isinstance(last_modified, datetime):
                meta["last_modified"] = last_modified.isoformat()
            else:
                meta["last_modified"] = str(last_modified)
        if not meta.get("seqno") and meta.get("etag"):
            meta["seqno"] = meta["etag"]
        return body, meta

    return await anyio.to_thread.run_sync(_get, limiter=None)


async def _download_ftp(parsed, timeout_s: int) -> tuple[bytes, dict[str, Any]]:
    path = parsed.path.lstrip("/")
    host = parsed.hostname
    if not host:
        raise ValueError("FTP source missing hostname")
    port = parsed.port or 21
    username = parsed.username or "anonymous"
    password = parsed.password or "anonymous@"

    def _get() -> tuple[bytes, dict[str, Any]]:
        from ftplib import FTP

        buf = io.BytesIO()
        meta: dict[str, Any] = {}
        with FTP() as ftp:
            ftp.connect(host, port, timeout=timeout_s)
            ftp.login(username, password)
            ftp.retrbinary(f"RETR {path}", buf.write)
            try:
                stat = ftp.sendcmd(f"MDTM {path}")
                if stat.startswith("213"):
                    timestamp = stat.split()[1]
                    meta["last_modified"] = timestamp
                    meta["seqno"] = timestamp
            except Exception:  # pragma: no cover - optional metadata
                pass
        return buf.getvalue(), meta

    return await anyio.to_thread.run_sync(_get, limiter=None)


def _parse_rows(source: str, raw: bytes, meta: dict[str, Any]) -> list[dict[str, Any]]:
    if not raw:
        return []
    hint = meta.get("content_type") or source
    fmt = _detect_format(raw, hint)
    if fmt == "csv":
        return _parse_csv_rows(source, raw)
    if fmt == "excel":
        return _parse_excel_rows(source, raw)
    diag = _format_diagnostics(source, meta, raw)
    raise UnsupportedFileFormatError(f"Unsupported data format for {diag}")


def _detect_format(raw_bytes: bytes, name_or_ct: str | None) -> str:
    hint = (name_or_ct or "").lower()
    if ".csv" in hint or "text/csv" in hint:
        return "csv"
    if any(ext in hint for ext in (".xlsx", ".xls", "spreadsheet", "ms-excel")):
        return "excel"
    # Fallback: try sniffing CSV
    try:
        sample = raw_bytes[:1024].decode("utf-8-sig")
        csv.Sniffer().sniff(sample)
        return "csv"
    except Exception:
        return "excel"


def _format_diagnostics(source: str, meta: dict[str, Any], raw: bytes) -> str:
    ct = meta.get("content_type") or "unknown"
    magic = " ".join(f"{b:02x}" for b in raw[:16])
    return f"{source} (content_type={ct}, magic={magic})"


def _parse_csv_rows(source: str, raw: bytes) -> list[dict[str, Any]]:
    text = raw.decode("utf-8-sig")
    buf = io.StringIO(text)
    try:
        dialect = csv.Sniffer().sniff(text[:1024])
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(buf, dialect=dialect)
    rows: list[dict[str, Any]] = []
    for row in reader:
        try:
            rows.append(_normalize_row(row, source))
        except ValueError as exc:
            record_etl_normalize_error("logistics_etl", "row_error")
            logger.warning("Skipping invalid row from %s: %s", source, exc)
    record_etl_rows_normalized("logistics_etl", len(rows))
    return rows


def _parse_excel_rows(source: str, raw: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore[import]
    except Exception as exc:
        raise UnsupportedExcelError("openpyxl is required to process Excel logistics sources") from exc

    stream = io.BytesIO(raw)
    wb = load_workbook(stream, read_only=True, data_only=True)
    ws = wb.active
    rows: list[dict[str, Any]] = []
    headers: list[str] = []
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        values = list(row)
        if idx == 0:
            headers = [str(v).strip() if v is not None else "" for v in values]
            continue
        data = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
        try:
            rows.append(_normalize_row(data, source))
        except ValueError as exc:
            record_etl_normalize_error("logistics_etl", "row_error")
            logger.warning("Skipping invalid Excel row from %s: %s", source, exc)
    record_etl_rows_normalized("logistics_etl", len(rows))
    return rows


def _normalize_row(row: dict[str, Any], source: str) -> dict[str, Any]:  # noqa: C901
    normalized = {((k or "").strip().lower()): v for k, v in row.items()}

    def _as_str(*keys: str) -> str:
        for key in keys:
            value = normalized.get(key)
            if value is None:
                continue
            if isinstance(value, str):
                value = value.strip()
            return str(value).strip()
        raise ValueError(f"Missing required field {keys} in source {source}")

    def _optional_str(*keys: str) -> str | None:
        try:
            raw_value = _as_str(*keys)
        except ValueError:
            return None
        return raw_value or None

    def _as_decimal(*keys: str) -> Decimal:
        for key in keys:
            value = normalized.get(key)
            if value is None or value == "":
                continue
            try:
                return parse_decimal(value)
            except (TypeError, ValueError) as exc:
                raise ValueError(f"Invalid decimal for {key}: {value}") from exc
        raise ValueError(f"Missing decimal field {keys} in source {source}")

    def _as_date(*keys: str) -> date | None:
        raw_value = _optional_str(*keys)
        if raw_value is None:
            return None
        try:
            return vendor_parse_date(raw_value)
        except ValueError as exc:
            raise ValueError(f"Invalid date value {raw_value} for keys {keys}") from exc

    carrier = _as_str("carrier", "carrier_name", "carrier code")
    origin = _as_str("origin", "origin_code", "from")
    dest = _as_str("dest", "destination", "destination_code", "to")
    service = _as_str("service", "service_level", "mode")
    eur_per_kg = _as_decimal("eur_per_kg", "rate", "price_per_kg")
    valid_from = _as_date("effective_from", "valid_from", "start_date")
    valid_to = _as_date("effective_to", "valid_to", "end_date")

    return RateRowModel(
        carrier=carrier,
        origin=origin,
        dest=dest,
        service=service,
        eur_per_kg=eur_per_kg,
        valid_from=valid_from,
        valid_to=valid_to,
        source=source,
    ).model_dump()
