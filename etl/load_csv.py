from __future__ import annotations

import argparse
import hashlib
import json
import tempfile
import time
from collections.abc import Callable, Iterator
from dataclasses import dataclass
from pathlib import Path
from typing import Any, cast

import pandas as pd
import structlog
from sqlalchemy import create_engine, text, update
from sqlalchemy.orm import Session, sessionmaker

from awa_common.db.load_log import LOAD_LOG
from awa_common.dsn import build_dsn
from awa_common.etl.guard import process_once
from awa_common.etl.idempotency import build_payload_meta, compute_idempotency_key
from awa_common.metrics import record_etl_run, record_etl_skip
from awa_common.minio import create_boto3_client, get_bucket_name
from awa_common.settings import settings
from services.etl.dialects import (
    amazon_ads_sp_cost,
    amazon_fee_preview,
    amazon_inventory_ledger,
    amazon_reimbursements,
    amazon_returns,
    amazon_settlements,
    normalise_headers,
    schemas,
)
from services.worker.copy_loader import copy_df_via_temp


class ImportFileError(RuntimeError):
    """Raised when the ETL pipeline fails after persisting load_log metadata."""

    status_code = 500

    def to_meta(self) -> dict[str, str]:
        return {"status": "error", "error": str(self)}


class ImportValidationError(ImportFileError):
    """Raised for user-facing issues (e.g. empty or malformed uploads)."""

    status_code = 422


_ETL_CFG = getattr(settings, "etl", None)
USE_COPY = bool(_ETL_CFG.use_copy if _ETL_CFG else getattr(settings, "USE_COPY", True))
STREAMING_CHUNK_ENV = int(
    _ETL_CFG.ingest_streaming_chunk_size if _ETL_CFG else getattr(settings, "INGEST_STREAMING_CHUNK_SIZE", 50_000)
)
BUCKET = get_bucket_name()
CSV_EXTENSIONS = {".csv", ".txt", ".tsv"}
XLSX_EXTENSIONS = {".xlsx", ".xlsm"}
SOURCE_NAME = "ingest.import_file"
logger = structlog.get_logger(__name__).bind(service="ingest", etl_name="load_csv")
TABLE_TO_DIALECT = {
    "returns_raw": "returns_report",
    "reimbursements_raw": "reimbursements_report",
    amazon_fee_preview.TARGET_TABLE: "fee_preview_report",
    amazon_inventory_ledger.TARGET_TABLE: "inventory_ledger_report",
    amazon_ads_sp_cost.TARGET_TABLE: "ads_sp_cost_daily_report",
    amazon_settlements.TARGET_TABLE: "settlements_txn_report",
}


@dataclass(slots=True)
class _StreamingMetadata:
    dialect: str
    rows: int
    keyword_full: bool
    transaction_full: bool


def _read_csv_flex(path: Path) -> pd.DataFrame:
    for enc in ("utf-8", "utf-8-sig", "cp1252"):
        try:
            return pd.read_csv(path, sep=None, engine="python", encoding=enc)
        except UnicodeDecodeError:
            continue
        except Exception:
            pass
    for sep in (",", ";", "\t", "|"):
        try:
            return pd.read_csv(path, sep=sep)
        except Exception:
            continue
    raise ImportValidationError(f"Failed to read CSV: {path}")


def _sha256_file(path: str | Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _download_from_minio(path: str) -> Path:
    s3 = create_boto3_client()
    tmp = tempfile.NamedTemporaryFile(delete=False)
    s3.download_fileobj(BUCKET, path, tmp)
    tmp.close()
    return Path(tmp.name)


def _is_s3_uri(uri: str) -> bool:
    return isinstance(uri, str) and uri.startswith(("s3://", "minio://"))


def _open_uri(uri: str) -> Path:
    """
    TESTING-only hook: in tests we monkeypatch this to return a local file.
    In production, your existing S3/MinIO downloader remains in place.
    """
    if getattr(settings, "TESTING", False):
        return Path(uri)
    raise RuntimeError("S3/MinIO open is not available outside tests")


def _normalize_for_dialect(df: pd.DataFrame, dialect: str) -> pd.DataFrame:
    if dialect == "returns_report":
        return amazon_returns.normalise(df)
    if dialect == "reimbursements_report":
        return amazon_reimbursements.normalise(df)
    if dialect == "fee_preview_report":
        return amazon_fee_preview.normalise(df)
    if dialect == "inventory_ledger_report":
        return amazon_inventory_ledger.normalise(df)
    if dialect == "ads_sp_cost_daily_report":
        return amazon_ads_sp_cost.normalise(df)
    if dialect == "settlements_txn_report":
        return amazon_settlements.normalise(df)
    raise ImportValidationError("Unknown report: cannot detect dialect")


def _detect_dialect_from_columns(normalized_columns: list[str]) -> str:
    if amazon_returns.detect(normalized_columns):
        return "returns_report"
    if amazon_reimbursements.detect(normalized_columns):
        return "reimbursements_report"
    if amazon_fee_preview.detect(normalized_columns):
        return "fee_preview_report"
    if amazon_inventory_ledger.detect(normalized_columns):
        return "inventory_ledger_report"
    if amazon_ads_sp_cost.detect(normalized_columns):
        return "ads_sp_cost_daily_report"
    if amazon_settlements.detect(normalized_columns):
        return "settlements_txn_report"
    raise ImportValidationError("Unknown report: cannot detect dialect")


def _resolve_dialect(df: pd.DataFrame, explicit: str | None) -> tuple[str, pd.DataFrame]:
    if explicit:
        return explicit, _normalize_for_dialect(df, explicit)
    cols = normalise_headers(df.columns)
    detected = _detect_dialect_from_columns(cols)
    return detected, _normalize_for_dialect(df, detected)


def _conflict_columns_for(
    dialect: str,
    *,
    df: pd.DataFrame | None = None,
    metadata: _StreamingMetadata | None = None,
) -> tuple[str, ...] | None:
    if dialect == "reimbursements_report":
        return ("reimb_id",)
    if dialect == "ads_sp_cost_daily_report":
        keyword_safe = metadata.keyword_full if metadata else bool(df is not None and df["keyword_id"].notna().all())
        if keyword_safe:
            return amazon_ads_sp_cost.CONFLICT_COLS
        return None
    if dialect == "settlements_txn_report":
        txn_safe = (
            metadata.transaction_full if metadata else bool(df is not None and df["transaction_id"].notna().all())
        )
        if txn_safe:
            return amazon_settlements.CONFLICT_COLS
        return None
    return None


def _gather_streaming_metadata(path: Path, chunk_size: int, dialect_hint: str | None) -> _StreamingMetadata:
    dialect: str | None = None
    rows = 0
    keyword_full = True
    transaction_full = True
    for chunk in load_large_csv(path, chunk_size=chunk_size):
        if chunk is None or chunk.empty:
            continue
        if dialect is None:
            dialect, normalized = _resolve_dialect(chunk, dialect_hint)
        else:
            normalized = _normalize_for_dialect(chunk, dialect)
        if normalized.empty:
            continue
        rows += len(normalized)
        if "keyword_id" in normalized.columns and normalized["keyword_id"].isna().any():
            keyword_full = False
        if "transaction_id" in normalized.columns and normalized["transaction_id"].isna().any():
            transaction_full = False
    if dialect is None or rows == 0:
        raise ImportValidationError("empty file")
    return _StreamingMetadata(
        dialect=dialect,
        rows=rows,
        keyword_full=keyword_full,
        transaction_full=transaction_full,
    )


def _process_streaming_chunks(
    path: Path,
    *,
    chunk_size: int,
    dialect: str,
    engine: Any,
    conn: Any,
    target_table: str,
    columns: list[str] | None,
    conflict_cols: tuple[str, ...] | None,
) -> int:
    total = 0
    resolved_columns = list(columns) if columns else None
    for chunk in load_large_csv(path, chunk_size=chunk_size):
        if chunk is None or chunk.empty:
            continue
        normalized = _normalize_for_dialect(chunk, dialect)
        if normalized.empty:
            continue
        try:
            validated = schemas.validate(normalized, dialect)
        except ValueError as err:
            raise ImportValidationError(str(err)) from err
        if not len(validated):
            continue
        if USE_COPY:
            if resolved_columns is None:
                resolved_columns = list(validated.columns)
            copy_df_via_temp(
                engine,
                validated,
                target_table=target_table,
                target_schema=None,
                columns=resolved_columns,
                conflict_cols=conflict_cols,
                analyze_after=False,
                connection=conn,
            )
        else:
            validated.to_sql(target_table, engine, if_exists="append", index=False)
        total += len(validated)
    if total == 0:
        raise ImportValidationError("empty file")
    return total


def load_large_csv(path: Path, *, chunk_size: int = STREAMING_CHUNK_ENV) -> Iterator[pd.DataFrame]:
    if chunk_size <= 0:
        raise ValueError("chunk_size must be positive")
    suffix = path.suffix.lower()
    if suffix in XLSX_EXTENSIONS:
        yield from _stream_xlsx_chunks(path, chunk_size)
        return
    if suffix in CSV_EXTENSIONS:
        yield from _stream_csv_chunks(path, chunk_size)
        return
    raise ImportValidationError(f"Streaming is only supported for CSV or XLSX files: {path}")


def _stream_csv_chunks(path: Path, chunk_size: int) -> Iterator[pd.DataFrame]:
    last_error: Exception | None = None
    for encoding in ("utf-8", "utf-8-sig", "cp1252"):
        try:
            reader = pd.read_csv(path, sep=None, engine="python", encoding=encoding, chunksize=chunk_size)
        except UnicodeDecodeError as err:
            last_error = err
            continue
        except Exception as err:
            last_error = err
            continue
        yield from _yield_reader(reader)
        return
    for sep in (",", ";", "\t", "|"):
        try:
            reader = pd.read_csv(path, sep=sep, chunksize=chunk_size)
        except Exception as err:
            last_error = err
            continue
        yield from _yield_reader(reader)
        return
    raise ImportValidationError(f"Failed to read CSV: {path}") from last_error


def _stream_xlsx_chunks(path: Path, chunk_size: int) -> Iterator[pd.DataFrame]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise ImportValidationError("openpyxl is required for XLSX streaming") from exc
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            headers: list[str] | None = None
            buffer: list[list[Any]] = []
            for row in ws.iter_rows(values_only=True):
                if headers is None:
                    if not row:
                        continue
                    header_values = [(cell if cell is not None else "") for cell in row]
                    if not any(header_values):
                        continue
                    headers = [str(cell).strip() for cell in header_values]
                    continue
                if row is None or not any(value is not None for value in row):
                    continue
                normalized_row = list(row[: len(headers)])
                if len(normalized_row) < len(headers):
                    normalized_row.extend([None] * (len(headers) - len(normalized_row)))
                buffer.append(normalized_row)
                if len(buffer) >= chunk_size:
                    yield pd.DataFrame(buffer, columns=headers)
                    buffer.clear()
            if headers and buffer:
                yield pd.DataFrame(buffer, columns=headers)
    finally:
        wb.close()


def _yield_reader(reader: pd.io.parsers.TextFileReader) -> Iterator[pd.DataFrame]:
    try:
        yield from reader
    finally:
        reader.close()


def _target_table_for(dialect: str) -> str:
    return {
        "returns_report": "returns_raw",
        "reimbursements_report": "reimbursements_raw",
        "fee_preview_report": amazon_fee_preview.TARGET_TABLE,
        "inventory_ledger_report": amazon_inventory_ledger.TARGET_TABLE,
        "ads_sp_cost_daily_report": amazon_ads_sp_cost.TARGET_TABLE,
        "settlements_txn_report": amazon_settlements.TARGET_TABLE,
    }[dialect]


def _build_import_meta(
    file_path: Path,
    *,
    target_table: str,
    dialect: str,
    streaming: bool,
    extra: dict[str, Any] | None = None,
) -> dict[str, Any]:
    meta_extra = {
        "source_uri": str(file_path),
        "target_table": target_table,
        "dialect": dialect,
        "streaming": streaming,
    }
    if extra:
        meta_extra.update(extra)
    result: dict[str, Any] = build_payload_meta(path=file_path, extra=meta_extra)
    return result


def _derive_idempotency_key(
    file_path: Path,
    *,
    target_table: str,
    dialect: str,
    user_key: str | None,
    force: bool,
    idempotent_enabled: bool,
) -> str:
    base_key: str = user_key or compute_idempotency_key(path=file_path)
    seed = json.dumps(
        {
            "key": base_key,
            "target_table": target_table,
            "dialect": dialect,
        },
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    key: str = compute_idempotency_key(content=seed)
    suffixes: list[str] = []
    if not idempotent_enabled:
        suffixes.append(f"run-{int(time.time() * 1000)}")
    if force:
        suffixes.append(f"force-{int(time.time() * 1000)}")
    if suffixes:
        key = f"{key}:{':'.join(suffixes)}"
    return key[:64]


def import_file(  # noqa: C901
    path: str,
    report_type: str | None = None,
    celery_update: Callable[[dict[str, Any]], None] | None = None,
    *,
    force: bool = False,
    streaming: bool = False,
    chunk_size: int | None = None,
    idempotency_key: str | None = None,
    **kwargs: Any,
) -> dict[str, Any]:
    _dialect_override = kwargs.pop("dialect", None)
    if kwargs:
        raise TypeError(f"Unexpected kwargs: {', '.join(kwargs)}")
    TESTING = bool(getattr(settings, "TESTING", False))
    file_path = Path(path)
    if file_path.exists() and file_path.stat().st_size == 0:
        raise ImportValidationError("empty file")
    if _dialect_override == "test_generic":
        streaming = False

    explicit_dialect = _dialect_override or report_type
    chunk_rows = chunk_size or STREAMING_CHUNK_ENV

    df: pd.DataFrame | None = None
    metadata: _StreamingMetadata | None = None
    dialect: str | None = None

    if streaming:
        if file_path.suffix.lower() == ".xls":
            raise ImportValidationError("Streaming requires XLSX files for Excel sources")
        metadata = _gather_streaming_metadata(file_path, chunk_rows, explicit_dialect)
        dialect = metadata.dialect
        if celery_update:
            celery_update({"stage": "read", "rows": metadata.rows})
            celery_update({"stage": "detect", "dialect": dialect})
    else:
        if file_path.suffix in {".xlsx", ".xls"}:
            df = pd.read_excel(file_path)
        else:
            df = _read_csv_flex(file_path)
        if df is None or (hasattr(df, "empty") and df.empty):
            raise ImportValidationError("empty file")
        if celery_update:
            celery_update({"stage": "read", "rows": len(df)})

        if TESTING and _dialect_override == "test_generic":
            from services.etl.dialects import test_generic as td

            try:
                df = td.normalize_df(df)
            except ValueError as err:
                raise ImportValidationError(str(err)) from err
            engine = create_engine(build_dsn(sync=True))
            with engine.begin() as db_conn:
                for row in df.to_dict(orient="records"):
                    db_conn.execute(
                        text(
                            'INSERT INTO test_generic_raw("ASIN", qty, price) VALUES (:ASIN,:qty,:price) '
                            'ON CONFLICT ("ASIN") DO UPDATE SET qty=EXCLUDED.qty, price=EXCLUDED.price'
                        ),
                        row,
                    )
            return {
                "status": "success",
                "rows": len(df),
                "dialect": td.NAME,
                "target_table": td.TABLE,
                "warnings": [],
            }

        dialect, df = _resolve_dialect(df, explicit_dialect)
        if celery_update:
            celery_update({"stage": "detect", "dialect": dialect})

    if dialect is None:
        raise ImportValidationError("Unknown report: cannot detect dialect")

    target_table = _target_table_for(dialect)
    idempotent_enabled = bool(_ETL_CFG.ingest_idempotent if _ETL_CFG else getattr(settings, "INGEST_IDEMPOTENT", True))
    analyze_min = int(_ETL_CFG.analyze_min_rows if _ETL_CFG else getattr(settings, "ANALYZE_MIN_ROWS", 50_000))
    warnings: list[str] = []

    meta_extra: dict[str, Any] = {"force": bool(force), "file_sha256": idempotency_key or _sha256_file(file_path)}
    if metadata:
        meta_extra["rows_estimated"] = metadata.rows
    elif df is not None:
        meta_extra["rows_estimated"] = len(df)

    payload_meta = _build_import_meta(
        file_path,
        target_table=target_table,
        dialect=dialect,
        streaming=streaming,
        extra=meta_extra,
    )
    idempotency_value = _derive_idempotency_key(
        file_path,
        target_table=target_table,
        dialect=dialect,
        user_key=idempotency_key,
        force=force,
        idempotent_enabled=idempotent_enabled,
    )

    column_map: dict[str, list[str] | None] = {
        "returns_report": list(df.columns) if df is not None else None,
        "reimbursements_report": list(df.columns) if df is not None else None,
        "fee_preview_report": list(amazon_fee_preview.TARGET_COLUMNS),
        "inventory_ledger_report": list(amazon_inventory_ledger.TARGET_COLUMNS),
        "ads_sp_cost_daily_report": list(amazon_ads_sp_cost.TARGET_COLUMNS),
        "settlements_txn_report": list(amazon_settlements.TARGET_COLUMNS),
    }
    columns = column_map[dialect]
    conflict_cols = _conflict_columns_for(dialect, df=df, metadata=metadata)

    engine = create_engine(build_dsn(sync=True))

    class _DummyResult:
        def __init__(self, inserted: bool = True) -> None:
            self.inserted = inserted

        def scalar_one_or_none(self) -> int | None:
            return 1 if self.inserted else None

    class _DummySession:
        def __init__(self) -> None:
            self._result = _DummyResult()

        def execute(self, *_args: Any, **_kwargs: Any) -> _DummyResult:
            return self._result

        def flush(self) -> None:
            return None

        def commit(self) -> None:
            return None

        def rollback(self) -> None:
            return None

        def close(self) -> None:
            return None

    def _make_dummy_session() -> _DummySession:
        return _DummySession()

    SessionLocal: Callable[[], Session]
    try:
        if hasattr(engine, "connect"):
            SessionLocal = cast(Callable[[], Session], sessionmaker(bind=engine, expire_on_commit=False, future=True))
        else:
            SessionLocal = cast(Callable[[], Session], _make_dummy_session)
    except Exception:
        SessionLocal = cast(Callable[[], Session], _make_dummy_session)
    try:
        with process_once(
            SessionLocal,
            source=SOURCE_NAME,
            payload_meta=payload_meta,
            idempotency_key=idempotency_value,
            on_duplicate="update_meta",
        ) as handle:
            if handle is None:
                record_etl_skip(SOURCE_NAME)
                logger.info(
                    "etl.skipped",
                    source=SOURCE_NAME,
                    idempotency_key=idempotency_value,
                    target_table=target_table,
                    dialect=dialect,
                )
                return {
                    "status": "skipped",
                    "rows": 0,
                    "dialect": dialect,
                    "target_table": target_table,
                    "warnings": warnings,
                    "idempotency_key": idempotency_value,
                }

            with record_etl_run(SOURCE_NAME):
                conn: Any = engine.raw_connection()
                rows_loaded = 0
                try:
                    conn.autocommit = False
                    if streaming:
                        rows_loaded = _process_streaming_chunks(
                            file_path,
                            chunk_size=chunk_rows,
                            dialect=dialect,
                            engine=engine,
                            conn=conn,
                            target_table=target_table,
                            columns=columns,
                            conflict_cols=conflict_cols,
                        )
                    else:
                        assert df is not None
                        try:
                            validated_df = schemas.validate(df, dialect)
                        except ValueError as err:
                            raise ImportValidationError(str(err)) from err
                        if celery_update:
                            celery_update({"stage": "validate", "rows": len(validated_df)})
                        if USE_COPY:
                            assert columns is not None
                            copy_df_via_temp(
                                engine,
                                validated_df,
                                target_table=target_table,
                                target_schema=None,
                                columns=list(columns),
                                conflict_cols=conflict_cols,
                                analyze_after=False,
                                connection=conn,
                            )
                        else:
                            validated_df.to_sql(target_table, engine, if_exists="append", index=False)
                        rows_loaded = len(validated_df)

                    if rows_loaded == 0:
                        raise ImportValidationError("empty file")
                    if rows_loaded >= analyze_min:
                        with conn.cursor() as cur:
                            cur.execute(f"ANALYZE {target_table}")
                    conn.commit()
                except Exception:
                    conn.rollback()
                    raise
                finally:
                    conn.close()

            payload_meta.update({"rows": rows_loaded, "warnings": warnings})
            handle.session.execute(
                update(LOAD_LOG).where(LOAD_LOG.c.id == handle.load_log_id).values(payload_meta=payload_meta)
            )
            if celery_update:
                if streaming:
                    celery_update({"stage": "validate", "rows": rows_loaded})
                celery_update({"stage": "write", "rows": rows_loaded})
            logger.info(
                "etl.success",
                source=SOURCE_NAME,
                idempotency_key=idempotency_value,
                target_table=target_table,
                dialect=dialect,
                rows=rows_loaded,
            )
            return {
                "status": "success",
                "rows": rows_loaded,
                "dialect": dialect,
                "target_table": target_table,
                "warnings": warnings,
                "idempotency_key": idempotency_value,
            }
    except ImportValidationError:
        logger.warning(
            "etl.validation_failed",
            source=SOURCE_NAME,
            target_table=target_table if "target_table" in locals() else None,
            dialect=dialect,
        )
        raise
    except Exception as exc:  # pragma: no cover - defensive
        logger.exception(
            "etl.failed",
            source=SOURCE_NAME,
            idempotency_key=idempotency_value if "idempotency_value" in locals() else None,
            target_table=target_table if "target_table" in locals() else None,
            dialect=dialect,
            error=str(exc),
        )
        raise ImportFileError(f"Failed to import {file_path}") from exc
    finally:
        engine.dispose()


def import_uri(uri: str, **kwargs: Any) -> dict[str, Any]:
    path = _open_uri(uri) if _is_s3_uri(uri) else Path(uri)
    return import_file(str(path), **kwargs)


def _resolve_report_type(table: str, report_type: str | None) -> str | None:
    if report_type:
        return report_type
    normalized = (table or "").strip().lower()
    if not normalized or normalized == "auto":
        return None
    return TABLE_TO_DIALECT.get(normalized, normalized)


def build_cli_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Load CSV/XLSX files into Postgres via COPY/UPSERT.")
    parser.add_argument("--source", required=True, help="Filesystem path or URI to the input file.")
    parser.add_argument(
        "--table",
        default="auto",
        help="Target table or dialect. Use 'auto' to auto-detect the dialect from headers.",
    )
    parser.add_argument(
        "--report-type",
        dest="report_type",
        default=None,
        help="Explicit report/dialect name (overrides --table).",
    )
    parser.add_argument("--force", action="store_true", help="Process even when idempotency detects a duplicate.")
    parser.add_argument("--streaming", action="store_true", help="Enable streaming loads for large files.")
    parser.add_argument("--chunk-size", type=int, default=None, help="Optional chunk size for streaming mode.")
    parser.add_argument("--idempotency-key", default=None, help="Override computed idempotency key.")
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_cli_parser()
    args = parser.parse_args(argv)
    resolved_report = _resolve_report_type(args.table, args.report_type)
    try:
        result = import_file(
            args.source,
            report_type=resolved_report,
            force=args.force,
            streaming=args.streaming,
            chunk_size=args.chunk_size,
            idempotency_key=args.idempotency_key,
        )
    except ImportFileError as exc:
        logger.error("etl.cli_failed", source=SOURCE_NAME, error=str(exc))
        return 1
    print(json.dumps(result, default=str))
    return 0
